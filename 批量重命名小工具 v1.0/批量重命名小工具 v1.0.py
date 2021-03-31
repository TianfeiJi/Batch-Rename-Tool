"""
    Author: JTF & HJC
    Date: 2020/5/7
"""
from tkinter import *
from tkinter import filedialog, messagebox, simpledialog, ttk
import openpyxl as xl
import os


class BatchRenameTool(Frame):
    def __init__(self,master=None):
        super().__init__(master)
        self.master = master
        self.pack()

        self.mainWindow()

    def selectExcelFile(self):
        """
            读取Excel文件，选择字段所在行数和用于定位的字段名
            将获取到的Excel文件信息输出到mainWindow
        """
        self.entry01.delete(0,END)
        self.Excelpath = filedialog.askopenfilename(title='选择Excel文件', filetypes=[('Excel', '*.xlsx')])
        self.entry01.insert(INSERT,self.Excelpath)
        # 加载EXcel表格
        wb = xl.load_workbook(self.Excelpath)
        # 获取第一个sheet
        sheet = wb[wb.sheetnames[0]]

        self.student_infos = []
        
        self.key_line = simpledialog.askinteger( '字段所在行数','请输入字段名所在行数', initialvalue=1)
        for row in range(self.key_line+1, sheet.max_row + 1):
            self.student_info = {}
            for col in range(1, sheet.max_column + 1):
                key = sheet.cell(self.key_line, col).value
                if key:  # 判断列名是否有效
                    value = str(sheet.cell(row, col).value)
                    self.student_info[key] = value
            self.student_infos.append(self.student_info)
        # 得到关键字列表
        self.info_keys = list(self.student_infos[0].keys())

        self.text01 = Text(self.master,width=40,height=4,bg='lightgray',bd=0)
        self.text01.place(x=120,y=53)
        self.text01.insert(INSERT,'应收：' + str(sheet.max_row-1) + ' 人\n')
        self.text01.insert(INSERT,'当前可替换字段有：')

        t = 0
        for key in self.info_keys:
            t += 1
            self.text01.insert(INSERT,str(key) + ' ')
            # 第一行只能显示4个，剩余的在下一行输出
            if t == 4:
                self.text01.insert(INSERT,'\n')

        # 默认为第一个字段
        self.find_key = self.info_keys[0]
        self.text01.insert(INSERT,' \n当前选择的定位字段： ' + self.find_key)

        # 弹出子窗口
        self.child01_root = Toplevel(self.master)
        self.child01_root.title('字段选择')
        self.child01_root.geometry('300x150+200+200')

        self.child01_label01 = Label(self.child01_root,text='当前可替换字段名有：')
        self.child01_label01.place(x=15,y=20)

        self.child01_text01 = Text(self.child01_root,width=20,height=1,bg='lightgray',bd=0)
        self.child01_text01.place(x=143,y=25)

        self.child01_text02 = Text(self.child01_root,width=38,height=1,bg='lightgray',bd=0)
        self.child01_text02.place(x=17,y=40)

        # 把当前可替换字段插入到子窗口中
        if len(self.info_keys) <= 4:
            for key in self.info_keys:
                self.child01_text01.insert(INSERT,str(key) + ' ')
        else:
            count = 0
            for key in self.info_keys:
                count += 1
                self.child01_text01.insert(INSERT,str(key) + ' ')
                if count == 4:
                    self.child01_text01.insert(INSERT,'\n')
                if count > 4:
                    self.child01_text02.insert(INSERT,str(key) + ' ')
                

        self.child01_label02 = Label(self.child01_root,text='请选择定位字段：')
        self.child01_label02.place(x=38,y=60)

        var_key_word = StringVar()

        self.keyChosen = ttk.Combobox(self.child01_root,width=5,textvariable=var_key_word,state="readonly")
        self.keyChosen['values'] = self.info_keys
        self.keyChosen.current(0)
        self.keyChosen.place(x=140,y=62)
        
        def getFindKey():
            self.find_key = var_key_word.get()

            index_row = 3 + int(len(self.info_keys)/4)
            index = str(index_row) + '.11'
            # 把row行11列后的数据删除
            self.text01.delete(index,END)
            self.text01.insert(INSERT,str(self.find_key))

        self.child01_btn01 = Button(self.child01_root,text='确定',width=5,height=1,command=getFindKey)
        self.child01_btn01.place(x=205,y=58)


    def selectFolder(self):
        """
            选择文件夹，将获取到的文件夹信息输出到mainWindow
        """
        self.entry02.delete(0,END)
        self.Folderpath = filedialog.askdirectory()
        self.entry02.insert(INSERT,self.Folderpath)

        self.file_names = os.listdir(self.Folderpath)
        if len(os.listdir(self.Folderpath)) == 0:
            messagebox.showerror('Error','文件夹为空！')
        
        self.text02 = Text(self.master,width=40,height=1,bg='lightgray',bd=0)
        self.text02.place(x=120,y=155)
        self.text02.insert(INSERT,'当前文件夹共有：' )
        self.text02.insert(INSERT,str(len(os.listdir(self.Folderpath))) + '份 文件')

        
    def closeThisWindow(self):
        root.destroy()


    def isLegal(self,file_name):
        """
            检测输入的命名格式是否合法
        """
        if self.find_key not in file_name:
            if messagebox.askquestion('警告！',f'您输入的命名格式中没有 {self.find_key} ！\n点击"是"继续执行，点击"否"重新输入！\n继续执行后的文件名称中将不包含 {self.find_key} ！') == 'no':
                return False

        if '&' + self.find_key + '&' in file_name:
            if self.find_key not in file_name.replace('&' + self.find_key + '&','__TEMP__'):
                messagebox.showerror('Error',f'您输入的命名格式中只有 &{self.find_key}& ！\n但是没有 {self.find_key} ！\n不能存在同格式同名称文件！\n')
                return False

        for item in '\\/:*?"<>|':
            if item in file_name:
                messagebox.showerror('Error','文件名中不能包含下列任何字符：\n\\ / : * ? " < > |')
                return False

        return True


    def check(self):
        """
            检查文件夹中符合规范的文件数量以及具体文件名
            检查文件夹中的文件是否含有重复的key_value
            检查文件夹中的文件是否含有key_value
        """
        submitted_list = []

        self.file_names = os.listdir(self.Folderpath)
        self.module_name = self.entry03.get()       
        if self.isLegal(self.module_name):
            # 计数为了检查是否有重复的key_value
            key_count = {}
            for stu_info in self.student_infos:
                key_value = stu_info[self.find_key]
                key_count[key_value] = 0

            for file_name in self.file_names:
                for stu_info in self.student_infos:
                    key_value = stu_info[self.find_key]
                    if key_value in file_name:
                        key_count[key_value] += 1


            # 不符合标准命名格式的列表
            nonstandard_list = []

            # 含有重复的stu_info[self.find_key]的列表
            repeated_list = []
            # 含有stu_info[self.find_key]的列表
            with_key_value_list = []
            # 不含有stu_info[self.find_key]的列表
            without_key_value_list = []

    
            for old_name in self.file_names: 
                for stu_info in self.student_infos:     
                    key_value = stu_info[self.find_key]
                    if key_value in old_name:
                        # 含有stu_info[self.find_key]的列表
                        with_key_value_list.append(old_name)
                        if key_count[key_value] == 1:
                            submitted_list.append(key_value)
                            new_name = self.module_name + os.path.splitext(old_name)[1]
                            for info_key in self.info_keys:
                                if  str('&' + info_key + '&') in new_name:
                                    temp_str = new_name.replace('&' + info_key + '&','__TEMP__')
                                    new_name = temp_str.replace(info_key, stu_info[info_key])
                                    new_name = new_name.replace('__TEMP__',info_key)
                                elif info_key in new_name:
                                    new_name = new_name.replace(info_key, stu_info[info_key])
                            # 统计不符合标准命名格式的列表
                            if old_name != new_name:   
                                nonstandard_list.append(old_name)
                                # 注意：不同于without_key_value_list, nonstandard_list中的元素是含有stu_info[self.find_key]的，只是和用户输入的标准命名格式不一致。
                            break
                        elif key_count[key_value] >= 2:
                            repeated_list.append(old_name)

            # 没有key_value的文件 = 已提交的文件 - 有key_value的文件
            without_key_value_list = list(set(self.file_names).difference(submitted_list,with_key_value_list))
            
            if len(nonstandard_list) != 0 or len(without_key_value_list) != 0 or len(repeated_list) != 0:
                self.child02_root = Toplevel(self.master)
                self.child02_root.title('检查结果')
                self.child02_root.geometry('380x300+200+200')

                self.child02_text01 = Text(self.child02_root,bd=0)
                self.child02_text01.pack()

                if len(nonstandard_list) != 0:
                    self.child02_text01.insert(INSERT,'以下 ' + str(len(nonstandard_list)) + ' 份文件的名称不符合您输入的命名格式:\n')  
                    # 有find_key, 但是不符合用户输入的标准命名格式
                    for item in nonstandard_list:
                        self.child02_text01.insert(INSERT,item + '\n')
                    self.child02_text01.insert(INSERT,'-' * 40 + '\n')

                if len(without_key_value_list) != 0:
                    self.child02_text01.insert(INSERT,'以下 ' + str(len(without_key_value_list)) + ' 份文件的 ' + self.find_key + ' 有误：\n')  
                    # “有误”指不含find_key
                    for item in without_key_value_list:
                        self.child02_text01.insert(INSERT,item + '\n')
                    self.child02_text01.insert(INSERT,'-' * 40 + '\n')

                if len(repeated_list) != 0:
                    self.child02_text01.insert(INSERT,'以下 ' + str(len(repeated_list)) + ' 份文件的 ' + self.find_key + ' 重复：\n')   
                    # find_key重复
                    for item in repeated_list:
                        self.child02_text01.insert(INSERT,item + '\n')
                    self.child02_text01.insert(INSERT,'-' * 40 + '\n')
            else:
                messagebox.showinfo('检查结果','当前文件夹中的文件都符合您输入的命名格式')

                        
    def rename(self):
        """
            批量重命名
        """
        submitted_list = []
        # 根据关键字段的值的长度进行排序，防止因为包含关系误判文件归属
        self.student_infos.sort(key=lambda x: len(x[self.find_key]), reverse=True)
        # 获取文件夹下的所有文件(夹)
        self.file_names = os.listdir(self.Folderpath)
        # 获取输入的模板名称
        self.module_name = self.entry03.get()   
        # 如果输入的命名格式合法    
        if self.isLegal(self.module_name):
            # 用于记录key_value出现的次数
            key_count = {}
            # 先默认设为0
            for stu_info in self.student_infos:
                key_value = stu_info[self.find_key]
                key_count[key_value] = 0
            # 出现一次就+1
            for file_name in self.file_names:
                for stu_info in self.student_infos:
                    key_value = stu_info[self.find_key]
                    if key_value in file_name:
                        key_count[key_value] += 1
                        
            # 遍历所有文件
            for filename in self.file_names:    
                # 遍历所有学生信息
                for stu_info in self.student_infos:     
                    key_value = stu_info[self.find_key]
                    if key_value in filename:
                        # 如果该key_value只出现一次(不处理有相同key_value的文件，例如，当遇到名字相同的学生时，不进行处理)
                        if key_count[key_value] == 1:
                            submitted_list.append(key_value)
                            # 原文件的路径
                            old_file=os.path.join(self.Folderpath, filename)
                            new_name = self.module_name
                            # 遍历所有字段名，并进行替换
                            for info_key in self.info_keys:
                                # 不替换&&中间的字段
                                if  '&' + info_key + '&' in new_name:
                                    # 先把 &字段& 变成一个临时的值
                                    temp_str = new_name.replace('&' + info_key + '&','__TEMP__')
                                    # 然后替换掉不带有&&的其他字段
                                    new_name = temp_str.replace(info_key, stu_info[info_key])
                                    # 再把临时值替换为& &中的字段
                                    new_name = new_name.replace('__TEMP__',info_key)
                                elif info_key in new_name:
                                    new_name = new_name.replace(info_key, stu_info[info_key])
                            # 重命名后的新文件的路径
                            new_file=os.path.join(self.Folderpath, new_name)  
                            # 重命名  
                            os.rename(old_file, new_file + os.path.splitext(filename)[1])   
                            # 成功执行一次后，break出去，执行下一次重命名
                            break
                            
            if len(submitted_list) == 0:
                messagebox.showerror('Error','文件夹中的文件不含 ' + str(self.find_key) + ' ！')        

            if len(os.listdir(self.Folderpath)) - len(submitted_list) != 0:
                messagebox.showinfo('执行结果','已成功重命名 ' + str(len(submitted_list)) + '份 文件！\n' + \
                    str(len(os.listdir(self.Folderpath)) - len(submitted_list)) + \
                        '份 文件保留原名称！\n' + '详情请点击 检查！')
            else: 
                messagebox.showinfo('执行成功','已成功重命名全部 ' + str(len(submitted_list)) + '份 文件！')   

            # 所有学生列表
            all_list = [student[self.find_key] for student in self.student_infos]
            # 通过求所有学生列表和已交学生列表，得到未交名单
            not_submit_list = list(set(all_list).difference(submitted_list))

            if len(not_submit_list) != 0:
                child02 = Toplevel(self.master)
                child02.title('未交名单')
                child02.geometry('300x150+300+450')

                child_text01 = Text(child02,bd=0)
                child_text01.pack()
                child_text01.insert(INSERT,'未交人数：' + str(len(not_submit_list)) + ' 人\n')
                child_text01.insert(INSERT,'未交名单：\n')

                i = 0
                for not_submit in not_submit_list:
                    i += 1
                    child_text01.insert(INSERT,str(i) + '、' + not_submit + '\n')  
    

    def mainWindow(self):
        """
            主界面
        """
        self.label01 = Label(self.master,text='请选择EXCEL文件:')
        self.label01.place(x=15,y=30)

        self.entry01 = Entry(self.master,bg='white',width=40)
        self.entry01.place(x=120,y=30)

        self.btn01 = Button(self.master,text='浏览',width=8,command=self.selectExcelFile)
        self.btn01.place(x=410,y=25)

        self.label02 = Label(self.master,text='请选择文件夹:')
        self.label02.place(x=39,y=110)

        self.entry02 = Entry(self.master,bg='white',width=40)
        self.entry02.place(x=120,y=110)

        self.btn02 = Button(self.master,text='浏览',width=8,command=self.selectFolder)
        self.btn02.place(x=410,y=105)

        self.label03 = Label(self.master,text='请输入命名格式:')
        self.label03.place(x=27,y=195)

        self.entry03 = Entry(self.master,bg='white',width=40)
        self.entry03.place(x=120,y=195)

        # 当鼠标左键单击命名格式输入框时，弹出提示按钮
        def Tips(event):
            def showTips():
                messagebox.showinfo('提示','您输入的命名格式中的字段会被替换为相应的值\n如果不想替换某字段名，你可以在该字段前后加上 &\n例如：&学号&：学号')
            self.tipbtn =  Button(self.master,text='提示',width=8,command=showTips)
            self.tipbtn.place(x=410,y=193)
        self.entry03.bind("<Button-1>",Tips)

        self.CheckButton = Button(self.master,text='检查',width=8,command=self.check)
        self.CheckButton.place(x=95,y=255)

        self.RenameButton = Button(self.master,text='重命名',width=9,command=self.rename)
        self.RenameButton.place(x=215,y=255)

        self.QuitButton = Button(self.master,text='退出',width=8,command=self.closeThisWindow)
        self.QuitButton.place(x=335,y=255)


if __name__ == "__main__":
    root = Tk()
    root.geometry("500x330+500+200")
    root.title('批量重命名小工具 v1.0')
    app = BatchRenameTool(master=root)
            
    root.mainloop() 
