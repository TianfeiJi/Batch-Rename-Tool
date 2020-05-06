"""
    Author:iTeaCode & QQ453613434
    Date:2020/4/29
    Doc:解决了可能出现的bug，推荐使用。但只可以处理"学号","姓名"这两个字段的替换
"""
import openpyxl as xl
import os
import time
import sys
# 以对话框形式打开文件时取消注释
# from tkinter import filedialog

#学号列表
id_number_list = []
#姓名列表
name_list = []
#已提交学生列表
submitted_list = []
#未提交学生列表
not_submitted_list = []
#学生信息列表，里面存储字典
student_info_list = []


#操作excel
def excel():
    Excelpath =  "花名册.xlsx"
    # Excelpath = filedialog.askopenfilename(title='选择Excel文件', filetypes=[('Excel', '*.xlsx')])

    wb = xl.load_workbook(Excelpath)
    sheet = wb[wb.sheetnames[0]]
    #获取2行2列开始到最大行最大列的数据
    for row in range(2,sheet.max_row+1):
        one_student = {}
        for col in range(2,sheet.max_column+1):
            key = sheet.cell(1,col).value
            value = sheet.cell(row,col).value
            one_student[key] = value

        student_info_list.append(one_student)
    # 根据关键字段的值的长度进行排序，防止因为包含关系误判文件归属
    student_info_list.sort(key = lambda student: len(str(student['姓名'])), reverse=True)
    print('应收：' + str(sheet.max_row-1) + ' 人')


#检查
def Check(file_name, is_legal=True):
    if '姓名' not in file_name and is_legal:
        print('警告！')
        print('您输入的命名格式中没有"姓名"！')
        print('这将造成不可逆的影响！如若忽视，请输入ignore\n')
        return False
    if '学号' not in file_name:
        print('警告!')
        print('不能存在同格式同名称文件！命名格式中至少要有"学号"！\n')
        return False
    # 检查文件名是否包含非法字符
    for item in '\\/:*?"<>|':
        if item in file_name:
            print('警告！')
            print('文件名中不能包含下列任何字符：')
            print('\\ / : * ? " < > |')
            return False
    return True


def rename():
    Folderpath = ".\作业"
    # Folderpath = filedialog.askdirectory()
    file_names = os.listdir(Folderpath)

    if len(os.listdir(Folderpath)) == 0:
        print('文件夹为空！')
        time.sleep(60)
        sys.exit(0)
    print('当前文件夹共有：' + str(len(os.listdir(Folderpath))) + '份 文件')
    print('-' * 40)

    is_legal = True
    module_name = input('请输入文件名：')
    while not Check(module_name, is_legal):
        re_input = input('请重新输入：' if is_legal else '(已忽略警告)请重新输入：')
        if re_input == 'ignore':
            is_legal = False
            module_name = input('(已忽略警告)请重新输入：')
        else:
            module_name = re_input

    print('\n正在进行批量重命名...\n')
    time.sleep(0.5)
    count = 0
    #遍历文件夹中的文件
    for f in file_names:
        #遍历姓名列表中的学生姓名
        for student in student_info_list:
            #判断姓名是否在文件名称中，如果不在就不更改文件名称
            if student['姓名'] in f:
                if student['姓名'] in submitted_list:
                    break
                #统计已提交列表
                submitted_list.append(student['姓名'])
                #统计已提交数量（带有学生名字的文件数量）
                count += 1
                #文件旧名称
                old_file=os.path.join(Folderpath, f)
                #由于replace函数不修改原str值，所以此处需要一个新变量存储
                new_name = module_name.replace('学号',str(student['学号'])).replace('姓名',str(student['姓名']))   
                #文件新名称
                new_file=os.path.join(Folderpath, new_name)
                #重命名
                os.rename(old_file, new_file + os.path.splitext(f)[1])
                break
    
    if len(submitted_list) == 0:
        print('文件夹中的文件不含学生名字！')
        time.sleep(60)
        sys.exit(0)

    if len(os.listdir(Folderpath)) - count != 0:
        print('已成功重命名 ' + str(count) + '份 文件！\n')
        print(str(len(os.listdir(Folderpath)) - count) + '份 文件保留原名称')
    else:
        print('已成功重命名全部 ' + str(count) + '份 文件！ ')
    print('-' * 40)

    #求学生列表和已交列表的差集,得到未提交学生列表
    name_list = [student['姓名'] for student in student_info_list]

    not_submitted_list = list(set(name_list).difference(submitted_list))
    if len(not_submitted_list) != 0:
        print('未交人数：' + str(len(not_submitted_list)) + ' 人')
        print('未交名单：', end='')

        i = 0
        for not_submitted in not_submitted_list:
            i += 1
            print(not_submitted, end='  ')
            #1行输出5人
            if i%5 == 0:
                print('\n' + ' ' * 10,end='')

if __name__ == "__main__":
    excel()
    rename()
    #生成exe文件时加上
    # print()
    # time.sleep(60)
