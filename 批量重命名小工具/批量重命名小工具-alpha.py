"""
    Author:iTeaCode
    Date:2020/4/24
    此版本只简单实现了功能，对于bug未作处理，不建议使用
"""

import openpyxl as xl
import os
import time
import sys

#学号列表
id_number_list = []
#姓名列表
name_list = []
#已提交学生列表
submitted_list = []


#操作excel
wb = xl.load_workbook("花名册.xlsx")
sheet = wb["Sheet1"]
for row in range(2,sheet.max_row+1):
    #B2开始为ID
    coor_id = 'B' + str(row)
    #C2开始为name
    coor_name = 'C' + str(row)

    id_number_list.append(sheet[coor_id].value)
    name_list.append(sheet[coor_name].value)
print('应收：' + str(sheet.max_row-1) + ' 人')

#批量重命名
path = ".\作业"
file_names = os.listdir(path)
if len(os.listdir(path)) == 0:
    print('文件夹为空！')
    time.sleep(60)
    sys.exit(0)

print('当前文件夹共有：' + str(len(os.listdir(path))) + '份 文件')
print('-' * 40)

module_name = input('请输入命名格式：')

print('\n正在进行批量重命名\n')
time.sleep(0.5)
count = 0
#遍历文件夹中的文件
for f in file_names:
    #遍历姓名列表中的学生姓名
    for student_name in name_list:
        #判断姓名是否在文件名称中，如果不在就不更改文件名称
        if student_name in f:
            #统计已提交列表
            submitted_list.append(student_name)
            #统计已提交数量（带有学生名字的文件数量）
            count += 1
            #获取该学生的索引，然后通过索引得到该学生的学号
            index = name_list.index(student_name)
            #文件旧名称
            old_file=os.path.join(path, f)
            #由于replace函数不修改原str值，所以此处需要一个新变量存储
            new_name = module_name.replace('学号',str(id_number_list[index])).replace('姓名',student_name)     
            #文件新名称
            new_file=os.path.join(path, new_name)
            #重命名
            os.rename(old_file, new_file + os.path.splitext(f)[1])

  
if len(submitted_list) == 0:
    print('文件夹中的文件不含学生名字！')
    time.sleep(60)
    sys.exit(0)

if len(os.listdir(path)) - count != 0:
    print('已成功重命名 ' + str(count) + '份 文件！\n')
    print(str(len(os.listdir(path)) - count) + '份 文件保留原名称')
else:
    print('已成功重命名全部 ' + str(count) + '份 文件！ ')
print('-' * 40)


#生成exe文件时加上
print()
time.sleep(60)
